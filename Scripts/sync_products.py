#!/usr/bin/env python3
"""Synchronizace produktů z exportu (XLSX/CSV) do Supabase tabulky `products`.

Přidá do databáze POUZE nové produkty (ty, jejichž `code` v DB ještě není).
Stávající řádky nemění a nemaže. Nové produkty se vkládají BEZ `show_weee`
(zůstává NULL), takže se aplikace při prvním tisku sama zeptá na WEEE ikonu.

Použití:
    # jen náhled – nic nezapisuje (výchozí)
    python3 Scripts/sync_products.py products_22.xlsx

    # reálné vložení nových produktů
    python3 Scripts/sync_products.py products_22.xlsx --apply

Zápisový klíč:
    Ke čtení stačí veřejný anon klíč (je zabudovaný). K zápisu (--apply) je
    potřeba klíč s právem INSERT. Předej ho přes proměnnou prostředí:
        export SUPABASE_KEY="<service_role nebo anon s INSERT policy>"
    Pokud SUPABASE_KEY není nastaven, zkusí se zápis anon klíčem – když to
    RLS nepovolí, skript to nahlásí (HTTP 401/403) a nic se nestane.

Vstupní XLSX musí mít v prvním listu hlavičku se sloupci:
    code, pairCode, name, ean   (jako export ze Shoptetu)
CSV musí mít hlavičku: code, ean, name, pair_code
"""

import argparse
import csv
import json
import os
import sys
import urllib.request
import urllib.error

SUPABASE_URL = "https://osinlzagjimyrzjpdxai.supabase.co/rest/v1/products"
ANON_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9zaW5semFnamlteXJ6anBkeGFpIiwicm9sZSI6"
    "ImFub24iLCJpYXQiOjE3ODE2MDUzMDcsImV4cCI6MjA5NzE4MTMwN30."
    "aWkcUv9jpwbqQ3fSHZ_damRGwSqxC_YtH3siySoMgq4"
)


def read_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("Chybí knihovna openpyxl. Nainstaluj: pip3 install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}
    for req in ("code", "name", "ean"):
        if req not in idx:
            sys.exit(f"Ve sloupcích XLSX chybí '{req}'. Nalezeno: {header[:15]}")
    ci, ni, ei = idx["code"], idx["name"], idx["ean"]
    pi = idx.get("pairCode")

    items = []
    for r in rows:
        code = r[ci]
        if code is None or str(code).strip() == "":
            continue
        pair = r[pi] if pi is not None else None
        items.append({
            "code": str(code).strip(),
            "ean": str(r[ei]).strip() if r[ei] not in (None, "") else "",
            "name": str(r[ni]).strip() if r[ni] is not None else "",
            "pair_code": str(pair).strip() if pair not in (None, "") else None,
        })
    return items


def read_csv(path):
    items = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            code = (row.get("code") or "").strip()
            if not code:
                continue
            pair = (row.get("pair_code") or "").strip()
            items.append({
                "code": code,
                "ean": (row.get("ean") or "").strip(),
                "name": (row.get("name") or "").strip(),
                "pair_code": pair or None,
            })
    return items


def fetch_existing_codes(key):
    codes = set()
    offset, batch = 0, 1000
    while True:
        url = f"{SUPABASE_URL}?select=code&order=code"
        req = urllib.request.Request(url)
        req.add_header("apikey", key)
        req.add_header("Authorization", f"Bearer {key}")
        req.add_header("Range", f"{offset}-{offset + batch - 1}")
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
        codes.update(str(p["code"]) for p in data)
        if len(data) < batch:
            break
        offset += batch
    return codes


def insert_rows(rows, key, chunk=200):
    inserted = 0
    for i in range(0, len(rows), chunk):
        part = rows[i:i + chunk]
        body = json.dumps(part, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(SUPABASE_URL, data=body, method="POST")
        req.add_header("apikey", key)
        req.add_header("Authorization", f"Bearer {key}")
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
        try:
            with urllib.request.urlopen(req) as resp:
                resp.read()
            inserted += len(part)
            print(f"  vloženo {inserted}/{len(rows)}")
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", "replace")
            sys.exit(
                f"\nZápis selhal (HTTP {e.code}). Odpověď: {detail}\n"
                "Pokud jde o oprávnění (401/403), nastav SUPABASE_KEY na klíč "
                "s právem INSERT (service_role) a spusť znovu."
            )
    return inserted


def main():
    ap = argparse.ArgumentParser(description="Přidá do Supabase jen nové produkty.")
    ap.add_argument("input", help="Cesta k .xlsx (Shoptet export) nebo .csv")
    ap.add_argument("--apply", action="store_true",
                    help="Skutečně vloží nové produkty (jinak jen náhled).")
    args = ap.parse_args()

    if args.input.lower().endswith(".csv"):
        items = read_csv(args.input)
    else:
        items = read_xlsx(args.input)

    seen, deduped = set(), []
    for it in items:
        if it["code"] in seen:
            continue
        seen.add(it["code"])
        deduped.append(it)
    items = deduped
    print(f"V souboru: {len(items)} produktů (unikátní code)")

    read_key = ANON_KEY
    existing = fetch_existing_codes(read_key)
    print(f"V databázi: {len(existing)} produktů")

    new = [it for it in items if it["code"] not in existing]
    print(f"NOVÉ (přibudou): {len(new)}")

    if not new:
        print("Nic k přidání – databáze je aktuální.")
        return

    print("\nUkázka nových produktů:")
    for it in new[:25]:
        print(f"  {it['code']}  ean={it['ean'] or '-':<15} {it['name'][:55]}")
    if len(new) > 25:
        print(f"  … a dalších {len(new) - 25}")

    if not args.apply:
        print("\n[NÁHLED] Nic se nezapsalo. Pro vložení spusť znovu s --apply")
        return

    write_key = os.environ.get("SUPABASE_KEY", ANON_KEY)
    print(f"\nVkládám {len(new)} nových produktů…")
    n = insert_rows(new, write_key)
    print(f"Hotovo. Vloženo {n} produktů.")


if __name__ == "__main__":
    main()
